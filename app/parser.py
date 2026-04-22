from __future__ import annotations

import hashlib
import re
import shutil
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
from xml.etree import ElementTree as ET

import openpyxl

from app.models import CourtEvent

DATE_PATTERN = re.compile(
    r'(?P<day>\d{1,2})\.(?P<month>\d{1,2})\.(?P<year>\d{2,4})'
    r'(?:\s*(?:года|г\.|г)?\s*(?:в)?\s*(?P<hour>\d{1,2}):(?P<minute>\d{2}))?',
    flags=re.IGNORECASE,
)

IGNORE_VALUES = {
    '-',
    'дата не указана. надо разобраться',
    'дата не указана',
    '',
}

ARGB_RE = re.compile(r'^[0-9A-Fa-f]{8}$')
RGB_RE = re.compile(r'^[0-9A-Fa-f]{6}$')


def normalize_text(value: object) -> str:
    if value is None:
        return ''
    text = str(value).replace('\xa0', ' ').replace('\t', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


class SpreadsheetParser:
    def __init__(self, sheet_name: str, timezone_name: str, default_event_time: str):
        self.sheet_name = sheet_name
        self.tz = ZoneInfo(timezone_name)
        self.default_event_time = default_event_time

    def parse(self, file_path: str | Path) -> list[CourtEvent]:
        workbook = self._load_workbook_safe(file_path)

        if self.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet {self.sheet_name!r} not found. Available: {workbook.sheetnames}")

        ws = workbook[self.sheet_name]
        events: list[CourtEvent] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            case_number = normalize_text(row[0] if len(row) > 0 else None) or None
            claimant = normalize_text(row[1] if len(row) > 1 else None) or None
            respondent = normalize_text(row[2] if len(row) > 2 else None) or None
            raw_hearing_date = normalize_text(row[3] if len(row) > 3 else None)
            status = normalize_text(row[4] if len(row) > 4 else None) or None
            raw_deadline_date = normalize_text(row[5] if len(row) > 5 else None)

            if not any([case_number, claimant, respondent, raw_hearing_date, status, raw_deadline_date]):
                continue

            hearing_event = self._build_event(
                row_idx=row_idx,
                event_type="hearing",
                source_field="Дата судебного заседания",
                case_number=case_number,
                claimant=claimant,
                respondent=respondent,
                raw_date=raw_hearing_date,
                status=status,
            )
            if hearing_event is not None:
                events.append(hearing_event)

            deadline_event = self._build_event(
                row_idx=row_idx,
                event_type="deadline",
                source_field="Срок до",
                case_number=case_number,
                claimant=claimant,
                respondent=respondent,
                raw_date=raw_deadline_date,
                status=status,
            )
            if deadline_event is not None:
                events.append(deadline_event)

        return events

    def _load_workbook_safe(self, file_path: str | Path):
        try:
            return openpyxl.load_workbook(file_path, data_only=True)
        except ValueError as exc:
            message = str(exc).lower()
            if "stylesheet" not in message and "argb" not in message and "colors" not in message:
                raise

            repaired_path = self._repair_xlsx_styles(file_path)
            return openpyxl.load_workbook(repaired_path, data_only=True)

    def _repair_xlsx_styles(self, file_path: str | Path) -> Path:
        src = Path(file_path)
        tmp_dir = Path(tempfile.mkdtemp(prefix="xlsx_repair_"))
        fixed_xlsx = tmp_dir / f"{src.stem}_fixed.xlsx"

        with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(fixed_xlsx, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == "xl/styles.xml":
                    try:
                        data = self._fix_styles_xml(data)
                    except Exception:
                        # Если styles.xml совсем сломан, выбрасываем все стили целиком.
                        # Для чтения значений ячеек они не нужны.
                        data = self._minimal_styles_xml()

                zout.writestr(item, data)

        return fixed_xlsx

    def _fix_styles_xml(self, xml_bytes: bytes) -> bytes:
        root = ET.fromstring(xml_bytes)

        for elem in root.iter():
            rgb = elem.attrib.get("rgb")
            if rgb is None:
                continue

            rgb = rgb.strip()
            if ARGB_RE.fullmatch(rgb):
                continue
            if RGB_RE.fullmatch(rgb):
                elem.set("rgb", f"FF{rgb.upper()}")
            else:
                # Невалидное значение заменяем на безопасное чёрное с полной альфой.
                elem.set("rgb", "FF000000")

        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def _minimal_styles_xml(self) -> bytes:
        xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1">
    <font>
      <sz val="11"/>
      <name val="Calibri"/>
      <family val="2"/>
    </font>
  </fonts>
  <fills count="2">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
  </fills>
  <borders count="1">
    <border><left/><right/><top/><bottom/><diagonal/></border>
  </borders>
  <cellStyleXfs count="1">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
  </cellStyleXfs>
  <cellXfs count="1">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
  </cellXfs>
  <cellStyles count="1">
    <cellStyle name="Normal" xfId="0" builtinId="0"/>
  </cellStyles>
</styleSheet>
"""
        return xml.encode("utf-8")

    def _build_event(
        self,
        *,
        row_idx: int,
        event_type: str,
        source_field: str,
        case_number: str | None,
        claimant: str | None,
        respondent: str | None,
        raw_date: str,
        status: str | None,
    ) -> CourtEvent | None:
        parsed = self._parse_datetime(raw_date, event_type=event_type)
        if parsed is None:
            return None

        event_dt, event_date_str, event_time_str = parsed
        seed = "|".join([
            event_type,
            source_field,
            case_number or "",
            claimant or "",
            respondent or "",
            event_dt.isoformat(),
            status or "",
            str(row_idx),
        ])
        event_uid = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:20]

        return CourtEvent(
            event_uid=event_uid,
            event_type=event_type,
            source_field=source_field,
            case_number=case_number,
            claimant=claimant,
            respondent=respondent,
            event_date=event_date_str,
            event_time=event_time_str,
            event_datetime=event_dt,
            date_raw=raw_date,
            status=status,
            source_row=row_idx,
        )

    def _parse_datetime(self, raw_value: str, *, event_type: str):
        if not raw_value:
            return None

        lowered = raw_value.strip().lower()
        if lowered in IGNORE_VALUES:
            return None

        match = DATE_PATTERN.search(raw_value)
        if not match:
            return None

        day = int(match.group("day"))
        month = int(match.group("month"))
        year = int(match.group("year"))
        if year < 100:
            year += 2000

        default_hour, default_minute = [int(x) for x in self.default_event_time.split(":", maxsplit=1)]

        if event_type == "deadline":
            fallback_hour, fallback_minute = 18, 0
        else:
            fallback_hour, fallback_minute = default_hour, default_minute

        hour = int(match.group("hour")) if match.group("hour") is not None else fallback_hour
        minute = int(match.group("minute")) if match.group("minute") is not None else fallback_minute

        dt = datetime(year, month, day, hour, minute, tzinfo=self.tz)
        return dt, dt.date().isoformat(), dt.strftime("%H:%M")